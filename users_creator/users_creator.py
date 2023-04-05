import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Method to create dictionary with all the users info parsed from users.xlsx
def users_parser():
    workbook = openpyxl.load_workbook('users.xlsx')
    sheet = workbook['Sheet1']
    users_dict = dict()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        idn, first_name, last_name, email = row
        users_dict[f'first_name_{idn}'] = first_name
        users_dict[f'last_name_{idn}'] = last_name
        users_dict[f'email_{idn}'] = email
    workbook.close()
    return users_dict


# Method to count users quantity in the users.xlsx:
def users_count():
    workbook = openpyxl.load_workbook('users.xlsx')
    sheet = workbook['Sheet1']
    count = 0
    for row in sheet.iter_rows(min_row=2):
        if any(cell.value for cell in row):
            count += 1
    workbook.close()
    return count


def user_creator():
    user_qty = users_count()
    for i in range(1, user_qty + 1):
        service = webdriver.chrome.service.Service(executable_path=ChromeDriverManager().install())
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        browser = webdriver.Chrome(service=service, options=options)
        browser.maximize_window()
        wait = WebDriverWait(browser, 20)
        f_name = users_parser()[f'first_name_{i}']
        l_name = users_parser()[f'last_name_{i}']
        email = users_parser()[f'email_{i}']

        browser.get("http://localhost:3000/")
        sign_wdw = browser.current_window_handle
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.TextLink__link'))).click()
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter First Name"]')))\
            .send_keys(f_name)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,
                                                     'input[placeholder="Enter Last Name (Optional)"]'))) \
            .send_keys(l_name)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter Your Email"]'))) \
            .send_keys(email)

        # wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter Your Password"]')))\
        #     .send_keys('Godfather1989!')

        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter Your Password"]')))\
            .send_keys('G')
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter Your Password"]'))) \
             .send_keys('odfather1')
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Enter Your Password"]'))) \
            .send_keys('989!')
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Confirm Password"]'))) \
            .send_keys('Godfather1989!')
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[type="submit"]'))).click()
        browser.switch_to.new_window('tab')
        browser.get("http://localhost:8025/")
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.msglist-message'))).click()
        link = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#preview-plain > a'))).text
        modified_link = link[:17]+'3000'+link[21:]
        browser.switch_to.window(sign_wdw)
        browser.get(modified_link)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.ant-modal-root')))
        browser.quit()
        print(f'{f_name} {l_name}: {email} user was successfully created')
    print(f'Totally {user_qty} users were created')


user_creator()
